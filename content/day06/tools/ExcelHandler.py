import openpyxl 
import os,sys
from pathlib import Path
from openpyxl.styles import PatternFill

PROJECT_PATH=Path(__file__).resolve().parent.parent
# 经过Path的处理再os.path.join需要str化一下
sys.path.insert(0,str(PROJECT_PATH))
from config.imports import *

DATA_FILE_PATH=os.path.join(PROJECT_PATH,'test_data',r'test_cases.xlsx')

# 定义的无参数的装饰器
def common_exception(func):
    @functools.wraps(func)
    def wrapper(*args,**kargs):
        try:
            return func(*args,**kargs)
        except Exception as e:
            # 可以在这里添加日志记录，为了便于观察
            print(f"[Error] {func.__name__} 执行失败: {e}")
            # func.__name__打印函数名字，e就是失败原因
            raise e
    return wrapper

# 定义的Excel工具类
class ExcelHandler(object):
    def __init__(self,file_path:str):
        # 此处是实例化的变量，实例化变量不用返回值
        self.file_path=file_path
        self.wb=None
        self.sheet=None

    @common_exception   # 定义是就会执行，进行装饰
    def create_wb_sheet(self,sheet_name:Union[str,int]=0):
        # 打开指定表格，没有规定是只读还是只写，那就既能读也能写
        self.wb=openpyxl.load_workbook(self.file_path)
        if isinstance(sheet_name,str):
            self.sheet=self.wb[sheet_name]
            # 类型判断
        elif isinstance(sheet_name,int):
            self.sheet=self.wb.worksheets[sheet_name]
        else:
            print('你输入的值不符合')
             # 抛出异常而不是直接退出
            raise ValueError(f'sheet_name 必须是 str 或 int，收到: {type(sheet_name)}')
            '''
            以下都是内置的，不用导入，使用结合raise直接抛出问题，该函数不会继续执行，上层有except会继续处理，没有except则会终止
                ValueError
                TyepError
                KeyError
                IndexError
                FileNotFound
                PermissionError
            '''
            # 其他不正常情况，不想处理，就直接退出。
            # 但是：❌ 在类库中直接退出进程是危险操作
            # 应该：抛出异常让调用方决定如何处理。
            # exit(1)
        # return self.wb 有self的实例变量不用return

    @common_exception
    # 对文件只是读取内容就不用再进行save
    def read_point_sheet_rows(self,sheet_name:Union[str,int]=0)->List[List]:
        self.create_wb_sheet(sheet_name)
        all_data=[]
        for row in self.sheet.iter_rows(min_row=2,max_col=7,values_only=True):
            # values_only直接返回值
            if  bool(row[0]):
                all_data.append(list(row))
        return all_data

    
    
    def cover_write_point_sheet(self,sheet_name,data:List[list]):
        '''使用字典映射：首先需要有一个字典，然后判断关键字是不是在字典key中（直接判断），如果在，则可以直接使用字典改变其值。记住，字典的数据是重要的数据（最终保存数据）'''
        try:
            # 创建重要的数据的字典：O(n)
            # 带有条件判断，在关键字部位None条件下，最外层的括号是类型转换
            data_dict={row[0]:row for row in data if row[0]}

            green_fill=PatternFill(start_color='00FF00',end_color='00FF00',fill_type='solid')
            red_fill=PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
            yellow_fill=PatternFill(start_color='FFD700',end_color='FFD700',fill_type='solid')
            # 创建填充颜色字典        
            map_color={
                'Pass':green_fill,
                'Error':red_fill,
                'Fail':yellow_fill
            }
            self.create_wb_sheet(sheet_name)
            row_num=self.sheet.max_row
            # 从第二行开始
            for i in range(2,row_num+1):
                case_id=self.sheet.cell(row=i,column=1).value
                if case_id in data_dict:
                    row=data_dict[case_id]
                    self.sheet.cell(row=i,column=5).value=row[4]
                    self.sheet.cell(row=i,column=6).value=row[5]
                    self.sheet.cell(row=i,column=7).value=row[6]
                    status=str(row[5]).capitalize()
                    if status in map_color:
                        self.sheet.cell(row=i,column=6).fill=map_color[status]
        except Exception as e:
            raise type(e)(f'没能成功数据匹配写入，原因如下：{e}')
        finally:
            self.wb.save(self.file_path)

        # try:
        #     green_fill=PatternFill(start_color='00FF00',end_color='00FF00',fill_type='solid')
        #     red_fill=PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
        #     yellow_fill=PatternFill(start_color='FFD700',end_color='FFD700',fill_type='solid')
        #     # 对单元格格式预先颜色格式设置
        #     self.create_wb_sheet(sheet_name)
        #     # 获取最大有效函数和列数，下表从1开始
        #     row_num=self.sheet.max_row
        #     # col_num=self.sheet.max_column
        #     for i in range(2,row_num+2):
        #         # 寻找指定单元格通过数字指定均从下表1开始，除了worksheets是从0开始
        #         excel_col1_cell=self.sheet.cell(row=i,column=1)
        #         # 获取值
        #         cell_value=excel_col1_cell.value
        #         # 获取当前单元格的行（数字）
        #         row_id=excel_col1_cell.row
        #         # 获取当前单元格的列（数字）
        #         col_id=excel_col1_cell.col_idx
        #         # excel_col1_cell.value=''  赋值
        #         for index,row_value in enumerate(data):
        #             # 通过表格的第一列的primary_key来进行匹配，匹配成则进行颜色的填充以及值得填补
        #             # print(cell_value,row_value[0])
        #             if cell_value == row_value[0]:
        #                 # print(1)
        #                 self.sheet.cell(row=i,column=5).value=row_value[4]
        #                 self.sheet.cell(row=i,column=6).value=row_value[5]
        #                 self.sheet.cell(row=i,column=7).value=row_value[6]
                        
        #                 status_value=str(row_value[5]).lower().capitalize()
        #                 # print(status_value)
        #                 if str(status_value)=='Pass':
        #                     self.sheet.cell(row=i,column=6).fill=green_fill
        #                 elif str(status_value)=='Fail':
        #                     self.sheet.cell(row=i,column=6).fill=yellow_fill
        #                 # break目的是为了不必要的循环，当前匹配，那么后面绝不会再匹配，也就没必要继续
        #                 elif str(status_value)=='Error':
        #                     self.sheet.cell(row=i,column=6).fill=red_fill
        #                 break
        # except Exception as e:
        #     raise e
        # finally:
        #     # 写表格记得保存进工作簿
        #     self.wb.save(self.file_path)


# 单独运行时，才会调用该脚本，其他时候只会用上面的内容
if __name__=='__main__':
    # wb=openpyxl.load_workbook(r"C:\Users\DIY7SGH\Downloads\0528-J6M_BoschInternal PD-DID & PD-RID List V9.1 (1).xlsx")
    # sheet=wb.worksheets[2]
    # for row in sheet.iter_rows(min_row=2,values_only=True):
    #     print(row)
    excel=ExcelHandler(DATA_FILE_PATH)
    # 数字从下标0开始
    data=excel.read_point_sheet_rows(0)
    print(data)
    # for i,value in enumerate(data):
    #     # print(value)
    #     if i%2==0 and i!=0:
    #         value[5]='Pass'
    #     elif i%2!=0 and i!=0:
    #         value[5]='Fail'
    # # print(data)
    # excel.cover_write_point_sheet(0,data)