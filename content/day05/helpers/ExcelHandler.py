'''
    openpyxl notes:
        想要指定sheet打开, 首先需要加载workbook (wb=openpxl.load_workbook(path)) , 有两种方式:
            1. 指定sheet_name: wb['Sheet2 (自动创建的默认名) ']
            2. 通过索引: wb.worsheets[0] 默认从0开始, 程序员习惯
            3. 默认打开是原有第一个sheet: sheet=wb.active; 更改默认sheet: wb.active=wb[1]
        显示全部列名: wb.sheetnames
        遍历数据: sheet.iter_rows(min_row=1,max_row=1,min_col=1,max_col=1,values_only=True) 下表从1开始, 只返回值，而不是单元格对象
        写入数据: sheet.cell(row=1,column=1,value='值') 精准控制单元格, 下表从1开始; 只有sheet是从0开始
        添加数据: sheet.append(row:[])
        删除行数据: sheet.delete_rows(1,sheet.max_rows) 从第一行到最大一行
        设置单元格颜色: Pattern.Fill(start_color='FFC7CE',end_color='FFC7CE',fiil_type='solid') 就是左右颜色纯度一样, 一个色儿
        保存文件: wb.save(path)
'''
import openpyxl as opx
from openpyxl.styles import PatternFill
import os,traceback
from typing import List,Union,Dict, Generator

class ExcelHandle():
    def __init__(self,file_path,sheet_name):
        self.base_dir=os.path.dirname(__file__)
        self.project_path=os.path.join(self.base_dir)

        self.data_file=file_path
        self.sheet_name=sheet_name
        self.wb=None
        self.sheet=None
    
    def open_pointed_sheet(self):
        try:
            # wb=opx.load_workbook(file,read_only=True)
            self.wb=opx.load_workbook(self.data_file)
            # 此处设置为只读模式更加节省内存。反正是读数据又不是写数据，只用得到读，然后取值就行。
            if isinstance(self.sheet_name, str):
                return self.wb, self.wb[self.sheet_name]
            # isinstance 用来判断类
            elif isinstance(self.sheet_name, int):
                # 修正：整数索引要用 worksheets
                return self.wb, self.wb.worksheets[int(self.sheet_name)]
            
        except Exception as e:
            # 工具函数和主函数不用traceback.print_exc()
            raise Exception(f':\n没有能成功打开{self.data_file}的sheet:{self.sheet_name}') from e 
    
    def read_excel_cases(self) -> Generator[Dict,None,None]:
    # Generator 类型注解第一个是yield类型，.send()送入类型，当yield用完返回None
        try:
            # 获取指定sheet内容
            self.wb,self.sheet=self.open_pointed_sheet()
            # 逐行遍历cases, 控制行数从第2行开始，列数截止在地7行
            for row in self.sheet.iter_rows(min_row=2,max_col=7,values_only=True):
                if bool(row[0]):  # 或 if any(row_data.values()): 检查全部字段  ，不要空
                    yield {'case_id':row[0],'case_name':row[1],'keyword':row[2],'expected':row[3],'actual':row[4],'status':row[5],'remark':row[6]}
                # yield生成器针对大文本数据，可以逐行读取，而不是以下全部
        except Exception as e:
            # 这个函数在这儿算作主要的工具函数，出错还是需要打印错误吧？
            traceback.print_exc()
            raise Exception('没能读到数据') from e
        finally:
            if self.wb:
                self.wb.close()
    
    def write_excel_results(self, datas: List[List]):
        """
        追加模式写入Excel：
        - 第一次调用：写入表头 + 数据
        - 后续调用：在已有数据后追加新行
        """
        try:
            self.wb, self.sheet = self.open_pointed_sheet()
            
            # 配置颜色
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # 判断是否是空表（没有数据，只有表头或完全空白）
            is_empty = self.sheet.max_row <= 1
            
            # 如果是空表，先写入表头
            if is_empty:
                headers = ['case_id', 'case_name', 'keyword', 'expected', 'actual', 'status', 'remark']
                self.sheet.append(headers)
            
            # 记录起始行号（用于设置颜色）
            start_row = self.sheet.max_row + 1
            
            # 追加数据
            for row_data in datas:
                self.sheet.append(row_data)
            
            # 设置颜色（针对刚追加的行）
            for i, row_data in enumerate(datas):
                current_row = start_row + i
                if len(row_data) > 5:
                    status = str(row_data[5]).capitalize()
                    cell = self.sheet.cell(row=current_row, column=6)  # 第6列是status
                    
                    if status == 'Fail':
                        cell.fill = red_fill
                    elif status == 'Pass':
                        cell.fill = green_fill
            
            # 保存（注意：追加模式也要保存）
            self.wb.save(self.data_file)
            print(f"成功追加 {len(datas)} 条数据，当前共 {self.sheet.max_row-1} 行数据（含表头）")
            
        except Exception as e:
            traceback.print_exc()
            raise Exception('数据追加失败') from e
        finally:
            if self.wb:
                self.wb.close()