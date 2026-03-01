'''
    openpyxl notes:
        想要指定sheet打开, 首先需要加载workbook (wb=openpxl.load_workbook(path)) , 有两种方式:
            1. 指定sheet_name: wb['Sheet2 (自动创建的默认名) ']
            2. 通过索引: wb.worsheets[0] 默认从0开始, 程序员习惯
            3. 默认打开是原有第一个sheet: sheet=wb.active; 更改默认sheet: wb.active=wb[1]
        显示全部列名: wb.sheetnames
        遍历数据: sheet.iter_rows(min_row=1,max_row=1,min_col=1,max_col=1,values_only=True) 下表从1开始, 只返回值，而不是单元格对象
        写入数据: sheet.cell(row=1,column=1,value='值') 精准控制单元格, 下表从1开始; 只有sheet是从0开始
        添加数据: sheet.append(row:[])
        删除行数据: sheet.delete_rows(1,sheet.max_rows) 从第一行到最大一行
        设置单元格颜色: Pattern.Fill(start_color='FFC7CE',end_color='FFC7CE',fiil_type='solid') 就是左右颜色纯度一样, 一个色儿
        保存文件: wb.save(path)
'''
import openpyxl as opx
from openpyxl.styles import PatternFill
import os,traceback
from typing import List,Union,Dict, Generator

def open_pointed_sheet(file,sheet_name:[str,int]):
    wb=None
    try:
        # wb=opx.load_workbook(file,read_only=True)
        wb=opx.load_workbook(file)
        # 此处设置为只读模式更加节省内存。反正是读数据又不是写数据，只用得到读，然后取值就行。
        if isinstance(sheet_name, str):
            return wb, wb[sheet_name]
        # isinstance 用来判断类
        elif isinstance(sheet_name, int):
            # 修正：整数索引要用 worksheets
            return wb, wb.worksheets[int(sheet_name)]
        
    except Exception as e:
        # 工具函数和主函数不用traceback.print_exc()
        raise Exception(f':\n没有能成功打开{file}的sheet:{sheet_name}') from e

def read_excel_cases(data_file:str,sheet_name:int |str) -> Generator[Dict,None,None]:
    # Generator 类型注解第一个是yield类型，.send()送入类型，当yield用完返回None
    wb=None
    try:
        # 获取指定sheet内容
        wb,sheet=open_pointed_sheet(data_file,sheet_name)
        # 逐行遍历cases, 控制行数从第2行开始，列数截止在地7行
        for row in sheet.iter_rows(min_row=2,max_col=7,values_only=True,max_row=sheet.max_row):
            yield {'case_id':row[0],'case_name':row[1],'keyword':row[2],'expected':row[3],'actual':row[4],'status':row[5],'remark':row[6]}
            # yield生成器针对大文本数据，可以逐行读取，而不是以下全部
    except Exception as e:
        # 这个函数在这儿算作主要的工具函数，出错还是需要打印错误吧？
        traceback.print_exc()
        raise Exception('没能读到数据') from e
    finally:
        if wb:
            wb.close()

def write_excel_results(data_file:str,sheet_name,datas:List[List]):
    wb=None
    try:
        # 获取指定sheet内容
        wb,sheet=open_pointed_sheet(data_file,sheet_name)
        # 配置单元格格式
        greenfill=PatternFill(start_color='C6EFCE',end_color='C6EFCE',fill_type='solid')
        redfill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid')
        headers = ['case_id', 'case_name', 'keyword', 'expected', 'actual', 'status', 'remark']
        sheet.append(headers)
        for row_idx,row_data in enumerate(datas):
            # start=2 表示数字从2开始计数
            sheet.append(row_data)
            # 需要先添加再设置格式
            if len(row_data) > 5 and str(row_data[5]).capitalize()=='Fail':
                # 使用数组注意长度
                sheet.cell(row=row_idx+2,column=6).fill=redfill
            elif len(row_data) > 5 and str(row_data[5]).capitalize()=='Pass':
                sheet.cell(row=row_idx+2,column=6).fill=greenfill
                # sheet.cell从下表1开始，去除headers就该是从2开始
            else:
                    pass
    except Exception as e:
        # 这个函数在这儿算作主要的工具函数，出错还是需要打印错误吧？
        traceback.print_exc()
        raise Exception('数据没能成功写入') from e
    finally:
        if wb:
            sheet.delete_rows(1, sheet.max_row)
            # 清除数据，在保存
            wb.save(data_file)
            # 直接保存覆盖
            wb.close()



def main():
    try:
        base_dir=os.path.dirname(__file__)
        # print(base_dir)
        # 控制切分方向和次数
        # project_path=base_dir.rsplit('\\',1)[0]
        # 跨平台问题
        project_path=os.path.join(base_dir)
        data_file=os.path.join(project_path,'data','test_cases.xlsx')
        # print(data_file)

        sheet_name=0
        rows=read_excel_cases(data_file,sheet_name)
        # 判断返回结果是不是None
        rows=(rows if rows else None)
        # 快速遍历打印
        for row in rows:
            print(row)
    except Exception as e:
        # 主函数就不用继续推卸则任了，需要对返回的错误进行分析统计
        print(f'当前错误是{e}')

if __name__=='__main__':
    main()